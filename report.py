"""
report.py  --  Poisson 1D Jacobi benchmark report generator.

Reads CSVs produced by benchmark.sh and writes an Excel workbook with
five analytical sheets:

  1. Serial        serial_std vs serial_opt vs serial_cache
  2. Hilos         serial_std vs threads(2,4,6,8)
  3. Procesos      serial_std vs processes(2,4,6,8)
  4. Correlacion   threads(p) vs processes(p) at same p and n (side by side)
  5. Comparacion   Best representative of each strategy

Each sheet (except Correlacion) contains:
  Table 1  -- individual measurements (reps as rows, grid sizes as cols)
              with Avg / StdDev / CV% and avg iters_done summary rows
  Table 2  -- per-impl average time + speedup over serial_std
  Two embedded charts (time and speedup)

CSV columns expected:
  suite, impl, parallelism, grid_size, repetition, wall_time_ms, iters_done
  (max_error column is loaded if present)

Usage:
    python report.py [results_dir]
    Default: results_dir = results/
"""

from __future__ import annotations

import math
import os
import sys
from dataclasses import dataclass

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from report_utils import (
    C, CHART_STYLE, FONT_NAME,
    Series,
    set_col_width,
    write_title_row, save_figure, plot_lines,
)

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

RESULTS_DIR = sys.argv[1].rstrip("/") if len(sys.argv) > 1 else "results"
OUTPUT_PATH = os.path.join(RESULTS_DIR, "reporte_poisson.xlsx")
CHARTS_DIR  = os.path.join(RESULTS_DIR, "charts")

SUITE_FILES = {
    "serial":    "data_serial.csv",
    "threads":   "data_threads.csv",
    "processes": "data_processes.csv",
}

IMPL_COLORS: dict[str, str] = {
    "serial_std":   "#FF6600",
    "serial_opt":   "#C00000",
    "serial_cache": "#2E75B6",
    "threads_2":    "#BBBBBB",
    "threads_4":    "#70AD47",
    "threads_6":    "#4472C4",
    "threads_8":    "#ED7D31",
    "processes_2":  "#BBBBBB",
    "processes_4":  "#5CB85C",
    "processes_6":  "#337AB7",
    "processes_8":  "#F0AD4E",
}

C_EXTRA: dict[str, str] = {
    "summary_bg": "EBF3FB",
    "summary_fg": "1F4E79",
    "impl_hdr":   "1F4E79",
    "sep":        "D9E1F2",
    "corr_t_bg":  "E2EFDA",
    "corr_p_bg":  "FCE4D6",
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------

# {Impl: {grid_size: [(rep, wall_time_ms, iters_done)]}}
AllReps = dict["Impl", dict[int, list[tuple[int, float, int]]]]


@dataclass(frozen=True)
class Impl:
    name:        str
    parallelism: int

    @property
    def label(self) -> str:
        labels = {
            "serial_std":   "Secuencial estándar",
            "serial_opt":   "Secuencial optimizado (-O3 full)",
            "serial_cache": "Secuencial alineado a cache",
        }
        if self.name in labels:
            return labels[self.name]
        if self.name == "threads":
            return f"Hilos Jacobi ({self.parallelism} hilos)"
        if self.name == "processes":
            return f"Procesos fork ({self.parallelism} procesos)"
        return f"{self.name} ({self.parallelism})"

    @property
    def short_label(self) -> str:
        if self.parallelism == 0:
            return self.name
        return f"{self.name}_{self.parallelism}"

    @property
    def color(self) -> str:
        return IMPL_COLORS.get(self.short_label, "#888888")


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_csv(suite: str) -> pd.DataFrame | None:
    path = os.path.join(RESULTS_DIR, SUITE_FILES[suite])
    if not os.path.exists(path):
        print(f"  [--]  {path} not found")
        return None
    df = pd.read_csv(path)
    df.columns         = [c.strip().lower() for c in df.columns]
    df["impl"]         = df["impl"].str.strip()
    df["parallelism"]  = df["parallelism"].astype(int)
    df["grid_size"]    = df["grid_size"].astype(int)
    df["repetition"]   = df["repetition"].astype(int)
    df["wall_time_ms"] = df["wall_time_ms"].astype(float)
    df["iters_done"]   = df["iters_done"].astype(int)
    if "max_error" in df.columns:
        df["max_error"] = df["max_error"].astype(float)
    print(f"  [OK]  {path}  ({len(df)} rows)")
    return df


def build_all_reps(frames: list[pd.DataFrame]) -> AllReps:
    combined = pd.concat(frames, ignore_index=True)
    result: AllReps = {}
    for (name, par), grp in combined.groupby(["impl", "parallelism"]):
        impl = Impl(name=str(name), parallelism=int(str(par)))
        result[impl] = {}
        for size, sub in grp.groupby("grid_size"):
            result[impl][int(str(size))] = sorted(
                [(int(r), float(t), int(it))
                 for r, t, it in zip(sub["repetition"],
                                     sub["wall_time_ms"],
                                     sub["iters_done"])],
                key=lambda x: x[0],
            )
    return result


def compute_avgs(all_reps: AllReps) -> dict[Impl, dict[int, float]]:
    return {
        impl: {s: sum(t for _, t, _ in pairs) / len(pairs)
               for s, pairs in sizes.items()}
        for impl, sizes in all_reps.items()
    }


def compute_avg_iters(all_reps: AllReps) -> dict[Impl, dict[int, float]]:
    return {
        impl: {s: sum(it for _, _, it in pairs) / len(pairs)
               for s, pairs in sizes.items()}
        for impl, sizes in all_reps.items()
    }


def best_parallel(avg_data: dict[Impl, dict[int, float]],
                  sizes: list[int], impl_name: str,
                  ref: Impl) -> Impl | None:
    ref_avgs = avg_data.get(ref, {})
    if not ref_avgs:
        return None
    best_impl, best_sp = None, 0.0
    for impl, times in avg_data.items():
        if impl.name != impl_name:
            continue
        sp_vals = [ref_avgs[s] / times[s]
                   for s in sizes
                   if s in times and s in ref_avgs and times[s] > 0]
        if sp_vals:
            sp = sum(sp_vals) / len(sp_vals)
            if sp > best_sp:
                best_sp, best_impl = sp, impl
    return best_impl


def size_label(n: int) -> str:
    if n >= 1_000_000 and n % 1_000_000 == 0:
        return f"{n // 1_000_000}M"
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M"
    if n >= 1_000 and n % 1_000 == 0:
        return f"{n // 1_000}K"
    return str(n)


# ---------------------------------------------------------------------------
# Cell styling helpers
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border() -> Border:
    thick = Side(style="medium", color="1F4E79")
    thin  = Side(style="thin",   color="BDD7EE")
    return Border(left=thick, right=thick, top=thin, bottom=thick)


def _hdr(cell, value: str, bg: str, fg: str = "FFFFFF",
         size: int = 10, bold: bool = True, align: str = "center") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=size, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=True)
    cell.border    = _thin_border()


def _dat(cell, value, fmt: str | None = None,
         bg: str = "FFFFFF", fg: str = "000000",
         bold: bool = False, align: str = "right") -> None:
    cell.value     = value
    cell.font      = Font(name=FONT_NAME, bold=bold, size=10, color=fg)
    cell.fill      = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _thin_border()
    if fmt:
        cell.number_format = fmt


def _summary_dat(cell, value, fmt: str | None = None,
                 bold: bool = False) -> None:
    _dat(cell, value, fmt=fmt,
         bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"], bold=bold)


# ---------------------------------------------------------------------------
# Chart generation
# ---------------------------------------------------------------------------

def chart_time(impls: list[Impl], avg_data: dict[Impl, dict[int, float]],
               sizes: list[int], title: str, fname: str) -> str:
    series = [
        Series(label=impl.short_label,
               data={s: avg_data[impl][s] for s in sizes
                     if s in avg_data.get(impl, {})},
               color=impl.color)
        for impl in impls if impl in avg_data
    ]
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        plot_lines(ax, [s for s in series if s.data], log_scale=True)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Tamaño de grilla (N)", fontsize=10)
        ax.set_ylabel("Tiempo promedio (ms)", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)


def chart_speedup(impls: list[Impl], avg_data: dict[Impl, dict[int, float]],
                  ref: Impl, sizes: list[int], title: str, fname: str) -> str:
    ref_avgs = avg_data.get(ref, {})
    series = []
    for impl in impls:
        if impl == ref or impl not in avg_data:
            continue
        data = {s: ref_avgs[s] / avg_data[impl][s]
                for s in sizes
                if s in avg_data[impl] and s in ref_avgs
                and avg_data[impl][s] > 0}
        if data:
            series.append(Series(label=impl.short_label,
                                 data=data, color=impl.color))
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--",
                   label=f"ref: {ref.short_label}")
        plot_lines(ax, series, log_scale=False)
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Tamaño de grilla (N)", fontsize=10)
        ax.set_ylabel("Speedup", fontsize=10)
        ax.legend(fontsize=8, loc="upper left")
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)


def chart_correlation(par_counts: list[int],
                      avg_data: dict[Impl, dict[int, float]],
                      ref: Impl, sizes: list[int],
                      title: str, fname: str) -> str:
    """Grouped bar chart: avg speedup of threads(p) vs processes(p) per p."""
    ref_avgs = avg_data.get(ref, {})
    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(9, 5))
        x     = list(range(len(par_counts)))
        width = 0.35
        for offset, impl_name, color, lbl in [
            (-width / 2, "threads",   "#4472C4", "threads"),
            ( width / 2, "processes", "#ED7D31", "processes"),
        ]:
            sp_vals = []
            for p in par_counts:
                impl = Impl(impl_name, p)
                t    = avg_data.get(impl, {})
                sp_list = [ref_avgs[s] / t[s]
                           for s in sizes
                           if s in t and s in ref_avgs and t[s] > 0]
                sp_vals.append(
                    sum(sp_list) / len(sp_list) if sp_list else 0)
            bars = ax.bar([xi + offset for xi in x], sp_vals,
                          width=width * 0.9, label=lbl, color=color,
                          edgecolor="white", alpha=0.85)
            for bar, val in zip(bars, sp_vals):
                if val > 0:
                    ax.text(bar.get_x() + bar.get_width() / 2,
                            bar.get_height() + 0.02,
                            f"{val:.2f}x", ha="center", va="bottom",
                            fontsize=8, fontweight="bold")
        ax.axhline(1, color="#AAAAAA", lw=1.2, ls="--", label="ref")
        ax.set_xticks(x)
        ax.set_xticklabels([f"p={p}" for p in par_counts])
        ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
        ax.set_xlabel("Número de trabajadores (p)", fontsize=10)
        ax.set_ylabel("Speedup promedio", fontsize=10)
        ax.legend(fontsize=9)
        fig.tight_layout()
    return save_figure(fig, CHARTS_DIR, fname)


# ---------------------------------------------------------------------------
# Standard sheet writer
# ---------------------------------------------------------------------------

def _write_raw_table(ws, impls: list[Impl], all_reps: AllReps,
                     avg_iters: dict[Impl, dict[int, float]],
                     sizes: list[int], n_cols: int,
                     start_row: int, n_reps: int) -> int:
    cur = start_row

    ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
    lbl = ws.cell(cur, 1, value="Tabla 1  —  Mediciones individuales (ms)")
    lbl.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    lbl.fill      = PatternFill("solid", fgColor=C["dark"])
    lbl.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    for impl_idx, impl in enumerate(impls):
        impl_reps = all_reps.get(impl, {})

        ws.merge_cells(f"A{cur}:{get_column_letter(n_cols)}{cur}")
        bh = ws.cell(cur, 1, value=impl.label)
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C_EXTRA["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        bh.border    = _thick_border()
        ws.row_dimensions[cur].height = 20
        cur += 1

        _hdr(ws.cell(cur, 1), "Rep", bg=C["mid"], size=9)
        for ci, size in enumerate(sizes, 2):
            _hdr(ws.cell(cur, ci), f"N = {size_label(size)}",
                 bg=C["mid"], size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for rep in range(1, n_reps + 1):
            bg = C["alt"] if rep % 2 == 0 else "FFFFFF"
            _dat(ws.cell(cur, 1), rep, fmt="0",
                 bg=C["light"], bold=True, align="center")
            for ci, size in enumerate(sizes, 2):
                val = next(
                    (t for r, t, _ in impl_reps.get(size, []) if r == rep),
                    None,
                )
                if val is not None:
                    _dat(ws.cell(cur, ci), round(val, 3),
                         fmt="#,##0.000", bg=bg)
                else:
                    _dat(ws.cell(cur, ci), "—", bg=bg, align="center")
            ws.row_dimensions[cur].height = 16
            cur += 1

        summary_defs = [
            ("Promedio",   "#,##0.000", 0),
            ("Desv. Est.", "#,##0.000", 1),
            ("CV (%)",     "0.00",      2),
        ]
        for s_label, fmt, s_idx in summary_defs:
            _hdr(ws.cell(cur, 1), s_label,
                 bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"],
                 size=9, align="left")
            for ci, size in enumerate(sizes, 2):
                vals = [t for _, t, _ in impl_reps.get(size, [])]
                if vals:
                    mean = sum(vals) / len(vals)
                    std  = math.sqrt(
                        sum((v - mean) ** 2 for v in vals) / len(vals))
                    cv   = std / mean * 100 if mean > 0 else 0.0
                    show = mean if s_idx == 0 else (std if s_idx == 1 else cv)
                    _summary_dat(ws.cell(cur, ci),
                                 round(show, 3 if s_idx < 2 else 2),
                                 fmt=fmt, bold=(s_idx == 0))
                else:
                    _summary_dat(ws.cell(cur, ci), "—")
            ws.row_dimensions[cur].height = 16
            cur += 1

        # Average iterations row
        _hdr(ws.cell(cur, 1), "Iter. prom.",
             bg=C_EXTRA["summary_bg"], fg=C_EXTRA["summary_fg"],
             size=9, align="left")
        for ci, size in enumerate(sizes, 2):
            it = avg_iters.get(impl, {}).get(size)
            if it is not None:
                _summary_dat(ws.cell(cur, ci),
                             int(round(it)), fmt="#,##0")
            else:
                _summary_dat(ws.cell(cur, ci), "—")
        ws.row_dimensions[cur].height = 16
        cur += 1

        if impl_idx < len(impls) - 1:
            for ci in range(1, n_cols + 1):
                ws.cell(cur, ci).fill = PatternFill(
                    "solid", fgColor=C_EXTRA["sep"])
            ws.row_dimensions[cur].height = 6
            cur += 1

    return cur


def _write_speedup_table(ws, impls: list[Impl],
                         avg_data: dict[Impl, dict[int, float]],
                         ref: Impl, sizes: list[int],
                         start_row: int) -> int:
    cur       = start_row
    n_sp_cols = 1 + len(sizes) * 2 + 1

    ws.merge_cells(f"A{cur}:{get_column_letter(n_sp_cols)}{cur}")
    sec = ws.cell(cur, 1,
                  value="Tabla 2  —  Promedio y Speedup  (ref: serial_std)")
    sec.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
    sec.fill      = PatternFill("solid", fgColor=C["dark"])
    sec.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[cur].height = 20
    cur += 1

    set_col_width(ws, 1, 32)
    for si in range(len(sizes)):
        set_col_width(ws, 2 + si * 2, 13)
        set_col_width(ws, 3 + si * 2, 10)
    set_col_width(ws, n_sp_cols, 12)

    _hdr(ws.cell(cur, 1), "Implementación", bg=C["dark"], size=10)
    for si, size in enumerate(sizes):
        _hdr(ws.cell(cur, 2 + si * 2),
             f"N={size_label(size)}\nAvg (ms)", bg=C["mid"], size=9)
        _hdr(ws.cell(cur, 3 + si * 2),
             f"N={size_label(size)}\nSpeedup",  bg=C["mid"], size=9)
    _hdr(ws.cell(cur, n_sp_cols), "Speedup\nProm.", bg=C["dark"], size=9)
    ws.row_dimensions[cur].height = 28
    cur += 1

    ref_avgs = avg_data.get(ref, {})

    for ri, impl in enumerate(impls):
        bg     = C["alt"] if ri % 2 == 0 else "FFFFFF"
        is_ref = (impl == ref)
        times  = avg_data.get(impl, {})

        _dat(ws.cell(cur, 1), impl.label,
             bg=C["light"], bold=True, align="left")

        sp_refs: list[str] = []
        for si, size in enumerate(sizes):
            ac, sc  = 2 + si * 2, 3 + si * 2
            avg_t   = times.get(size)
            ref_avg = ref_avgs.get(size)

            avg_c               = ws.cell(cur, ac)
            avg_c.value         = round(avg_t, 3) if avg_t is not None else "N/A"
            avg_c.number_format = "#,##0.000"
            avg_c.font          = Font(name=FONT_NAME, size=10)
            avg_c.fill          = PatternFill("solid", fgColor=bg)
            avg_c.alignment     = Alignment(horizontal="right",
                                            vertical="center")
            avg_c.border        = _thin_border()

            sp_c = ws.cell(cur, sc)
            if is_ref:
                sp_c.value = 1.0
            elif avg_t and avg_t > 0 and ref_avg:
                sp_c.value = round(ref_avg / avg_t, 4)
            else:
                sp_c.value = "N/A"
            sp_c.number_format = "0.0000"
            sp_c.font      = Font(name=FONT_NAME, size=10,
                                  color=C["green_fg"])
            sp_c.fill      = PatternFill("solid", fgColor=C["green_bg"])
            sp_c.alignment = Alignment(horizontal="right", vertical="center")
            sp_c.border    = _thin_border()

            if not is_ref:
                sp_refs.append(f"{get_column_letter(sc)}{cur}")

        avsp           = ws.cell(cur, n_sp_cols)
        avsp.value     = (f"=IFERROR(AVERAGE({','.join(sp_refs)}),\"N/A\")"
                          if sp_refs else 1.0)
        avsp.number_format = "0.00"
        avsp.font      = Font(name=FONT_NAME, bold=True, size=10,
                              color=C["green_fg"])
        avsp.fill      = PatternFill("solid", fgColor=C["green_bg"])
        avsp.alignment = Alignment(horizontal="right", vertical="center")
        avsp.border    = _thin_border()
        ws.row_dimensions[cur].height = 17
        cur += 1

    return cur


def write_sheet(wb: Workbook, sheet_name: str, title: str,
                impls: list[Impl], all_reps: AllReps,
                avg_iters: dict[Impl, dict[int, float]],
                ref: Impl, sizes: list[int],
                chart_time_path: str, chart_sp_path: str) -> None:
    ws       = wb.create_sheet(sheet_name)
    avg_data = compute_avgs(all_reps)
    n_reps   = max(
        (len(pairs) for impl in impls
         for pairs in all_reps.get(impl, {}).values()),
        default=5,
    )
    n_raw_cols = 1 + len(sizes)

    write_title_row(ws, title, n_raw_cols)
    ws.row_dimensions[1].height = 26

    set_col_width(ws, 1, 8)
    for ci in range(2, n_raw_cols + 1):
        set_col_width(ws, ci, 14)

    cur = _write_raw_table(ws, impls, all_reps, avg_iters, sizes,
                           n_raw_cols, start_row=2, n_reps=n_reps)
    cur += 2
    cur = _write_speedup_table(ws, impls, avg_data, ref, sizes,
                               start_row=cur)

    chart_anchor = cur + 3
    for anchor, path in [
        (f"A{chart_anchor}", chart_time_path),
        (f"L{chart_anchor}", chart_sp_path),
    ]:
        if path and os.path.exists(path):
            img        = XLImage(path)
            img.width  = 620
            img.height = 360
            ws.add_image(img, anchor)


# ---------------------------------------------------------------------------
# Correlation sheet: threads(p) vs processes(p) side by side
# ---------------------------------------------------------------------------

def write_correlation_sheet(wb: Workbook, all_reps: AllReps,
                             avg_iters: dict[Impl, dict[int, float]],
                             ref: Impl, par_counts: list[int],
                             sizes: list[int],
                             chart_path: str) -> None:
    ws       = wb.create_sheet("4. Correlacion")
    avg_data = compute_avgs(all_reps)
    ref_avgs = avg_data.get(ref, {})

    title = ("Correlación  |  threads(p) vs processes(p)  "
             "|  Speedup = T(serial_std) / T(impl)")
    # p | threads ms | threads sp | threads iters | processes ms | processes sp | processes iters
    N_BLOCK_COLS = 7
    write_title_row(ws, title, N_BLOCK_COLS)
    ws.row_dimensions[1].height = 26

    set_col_width(ws, 1, 8)
    for ci in range(2, N_BLOCK_COLS + 1):
        set_col_width(ws, ci, 15)

    cur = 2
    for size in sizes:
        ref_t = ref_avgs.get(size)
        ref_label = (f"serial_std avg: {ref_t:,.1f} ms"
                     if ref_t else "serial_std: —")

        ws.merge_cells(f"A{cur}:{get_column_letter(N_BLOCK_COLS)}{cur}")
        bh = ws.cell(cur, 1,
                     value=f"N = {size_label(size)}  |  {ref_label}")
        bh.font      = Font(name=FONT_NAME, bold=True, size=11, color="FFFFFF")
        bh.fill      = PatternFill("solid", fgColor=C_EXTRA["impl_hdr"])
        bh.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=1)
        bh.border    = _thick_border()
        ws.row_dimensions[cur].height = 20
        cur += 1

        sub_hdrs = [
            ("p",               C["dark"],          "FFFFFF"),
            ("Hilos (ms)",      C_EXTRA["corr_t_bg"], "1F4E79"),
            ("Hilos sp",        C_EXTRA["corr_t_bg"], "1F4E79"),
            ("Hilos iters",     C_EXTRA["corr_t_bg"], "1F4E79"),
            ("Procesos (ms)",   C_EXTRA["corr_p_bg"], "843C0C"),
            ("Procesos sp",     C_EXTRA["corr_p_bg"], "843C0C"),
            ("Procesos iters",  C_EXTRA["corr_p_bg"], "843C0C"),
        ]
        for ci, (text, bg, fg) in enumerate(sub_hdrs, 1):
            _hdr(ws.cell(cur, ci), text, bg=bg, fg=fg, size=9)
        ws.row_dimensions[cur].height = 18
        cur += 1

        for pi, p in enumerate(par_counts):
            bg     = C["alt"] if pi % 2 == 0 else "FFFFFF"
            t_impl = Impl("threads",   p)
            p_impl = Impl("processes", p)

            t_avg = avg_data.get(t_impl,  {}).get(size)
            p_avg = avg_data.get(p_impl,  {}).get(size)
            t_it  = avg_iters.get(t_impl, {}).get(size)
            p_it  = avg_iters.get(p_impl, {}).get(size)
            t_sp  = (ref_t / t_avg) if (t_avg and ref_t) else None
            p_sp  = (ref_t / p_avg) if (p_avg and ref_t) else None

            _dat(ws.cell(cur, 1), p, fmt="0",
                 bg=C["light"], bold=True, align="center")

            for col, val, fmt, color_bg, is_sp in [
                (2, t_avg, "#,##0.000",  C_EXTRA["corr_t_bg"], False),
                (3, t_sp,  "0.0000",     C_EXTRA["corr_t_bg"], True),
                (4, t_it,  "#,##0",      C_EXTRA["corr_t_bg"], False),
                (5, p_avg, "#,##0.000",  C_EXTRA["corr_p_bg"], False),
                (6, p_sp,  "0.0000",     C_EXTRA["corr_p_bg"], True),
                (7, p_it,  "#,##0",      C_EXTRA["corr_p_bg"], False),
            ]:
                cell = ws.cell(cur, col)
                if val is not None:
                    display = (int(round(val))
                               if isinstance(val, float) and fmt == "#,##0"
                               else round(val, 4)
                               if isinstance(val, float)
                               else val)
                    fg_color = C["green_fg"] if is_sp else "000000"
                    _dat(cell, display, fmt=fmt,
                         bg=color_bg, fg=fg_color, align="right")
                else:
                    _dat(cell, "—", bg=color_bg, align="center")

            ws.row_dimensions[cur].height = 16
            cur += 1

        for ci in range(1, N_BLOCK_COLS + 1):
            ws.cell(cur, ci).fill = PatternFill(
                "solid", fgColor=C_EXTRA["sep"])
        ws.row_dimensions[cur].height = 8
        cur += 1

    cur += 2
    if chart_path and os.path.exists(chart_path):
        img        = XLImage(chart_path)
        img.width  = 700
        img.height = 400
        ws.add_image(img, f"A{cur}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    os.makedirs(CHARTS_DIR, exist_ok=True)

    print("Reading CSVs...")
    df_serial    = load_csv("serial")
    df_threads   = load_csv("threads")
    df_processes = load_csv("processes")

    frames = [df for df in [df_serial, df_threads, df_processes]
              if df is not None]
    if not frames:
        print("No data found. Run benchmark.sh first.")
        sys.exit(1)

    all_reps  = build_all_reps(frames)
    avg_data  = compute_avgs(all_reps)
    avg_iters = compute_avg_iters(all_reps)
    sizes     = sorted({s for szs in all_reps.values() for s in szs})

    REF = Impl("serial_std", 0)

    def present(impl: Impl) -> bool:
        return impl in all_reps

    par_counts     = sorted({i.parallelism for i in all_reps
                              if i.parallelism > 0})
    thread_counts  = sorted({i.parallelism for i in all_reps
                              if i.name == "threads"})
    process_counts = sorted({i.parallelism for i in all_reps
                              if i.name == "processes"})

    proc_sizes = sorted({s for i, szs in all_reps.items()
                          if i.name == "processes" for s in szs})

    impls_serial    = [i for i in [REF,
                                    Impl("serial_opt",   0),
                                    Impl("serial_cache", 0)]
                       if present(i)]
    impls_threads   = [REF] + [Impl("threads",   t) for t in thread_counts
                                if present(Impl("threads",   t))]
    impls_processes = [REF] + [Impl("processes", p) for p in process_counts
                                if present(Impl("processes", p))]

    best_t = best_parallel(avg_data, sizes,      "threads",   REF)
    best_p = best_parallel(avg_data, proc_sizes, "processes", REF)
    impls_final = [i for i in [
        REF,
        Impl("serial_opt",   0),
        Impl("serial_cache", 0),
        best_t,
        best_p,
    ] if i is not None and present(i)]

    final_sizes = proc_sizes if best_p else sizes

    print("\nGenerating charts...")

    def gen(prefix: str, impls: list[Impl],
            chart_sizes: list[int],
            t_title: str, s_title: str) -> tuple[str, str]:
        ct = chart_time(impls, avg_data, chart_sizes,
                        t_title, f"{prefix}_time.png")
        cs = chart_speedup(impls, avg_data, REF, chart_sizes,
                           s_title, f"{prefix}_sp.png")
        print(f"  {os.path.basename(ct)}  |  {os.path.basename(cs)}")
        return ct, cs

    ct1, cs1 = gen("serial", impls_serial, sizes,
                   "Tiempo  |  Variantes seriales",
                   "Speedup  |  T(serial_std) / T(impl)")
    ct2, cs2 = gen("hilos", impls_threads, sizes,
                   "Tiempo  |  serial_std vs threads(p)",
                   "Speedup  |  T(serial_std) / T(threads, p)")
    ct3, cs3 = gen("procesos", impls_processes, proc_sizes,
                   "Tiempo  |  serial_std vs processes(p)",
                   "Speedup  |  T(serial_std) / T(processes, p)")

    corr_chart = chart_correlation(
        par_counts, avg_data, REF, proc_sizes,
        "Speedup promedio  |  threads(p) vs processes(p)  |  ref = serial_std",
        "correlacion_sp.png",
    )
    print(f"  {os.path.basename(corr_chart)}")

    ct5, cs5 = gen("final", impls_final, final_sizes,
                   "Tiempo  |  Comparación final por estrategia",
                   "Speedup  |  Mejor de cada estrategia  |  ref = serial_std")

    print("\nBuilding workbook...")
    wb = Workbook()
    if (default := wb.active) is not None:
        wb.remove(default)

    write_sheet(wb, "1. Serial",
                "Serial  |  serial_std  vs  serial_opt  vs  serial_cache",
                impls_serial, all_reps, avg_iters, REF, sizes, ct1, cs1)

    write_sheet(wb, "2. Hilos",
                "Hilos  |  serial_std  vs  threads(p)  |  ref = serial_std",
                impls_threads, all_reps, avg_iters, REF, sizes, ct2, cs2)

    write_sheet(wb, "3. Procesos",
                "Procesos  |  serial_std  vs  processes(p)  |  ref = serial_std",
                impls_processes, all_reps, avg_iters, REF,
                proc_sizes, ct3, cs3)

    write_correlation_sheet(
        wb, all_reps, avg_iters, REF,
        par_counts, proc_sizes, corr_chart,
    )

    write_sheet(wb, "5. Comparacion final",
                "Comparación final  |  Mejor de cada estrategia  |  ref = serial_std",
                impls_final, all_reps, avg_iters, REF,
                final_sizes, ct5, cs5)

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()